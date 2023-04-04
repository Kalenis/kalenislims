# -*- coding: utf-8 -*-
# This file is part of lims_interface module for Tryton.
# The COPYRIGHT file at the top level of this repository contains
# the full copyright notices and license terms.
import re
import sql
import formulas
import schedula
import unidecode
import io
import csv
import hashlib
import tempfile
import json
from openpyxl import load_workbook
from decimal import Decimal
from datetime import datetime, date, time
from dateutil import relativedelta
from itertools import chain
from collections import defaultdict

from trytond.config import config
from trytond.model import (Workflow, ModelView, ModelSQL, fields,
    sequence_ordered, Unique)
from trytond.wizard import (Wizard, StateTransition, StateView, StateAction,
    Button)
from trytond.pool import Pool
from trytond.pyson import PYSONDecoder, PYSONEncoder, Eval, Bool, Not, And, Or
from trytond.transaction import Transaction
from trytond.i18n import gettext
from trytond.exceptions import UserError
from .function import custom_functions

FUNCTIONS = formulas.get_functions()
FUNCTIONS.update(custom_functions)

FIELD_TYPES = [
    ('char', 'char', 'Text (single-line)', 'fields.Char', 'VARCHAR', str,
        None),
    ('multiline', 'text', 'Text (multi-line)', 'fields.Text', 'VARCHAR',
        str, None),
    ('integer', 'integer', 'Integer', 'fields.Integer', 'INTEGER', int, None),
    ('float', 'float', 'Float', 'fields.Float', 'FLOAT', float, None),
    ('numeric', 'numeric', 'Numeric', 'fields.Numeric', 'NUMERIC', Decimal,
        None),
    ('boolean', 'boolean', 'Boolean', 'fields.Boolean', 'BOOLEAN', bool,
        None),
    ('many2one', 'many2one', 'Link To Kalenis', 'fields.Many2One', 'INTEGER',
        int, None),
    ('date', 'date', 'Date', 'fields.Date', 'DATE', date, None),
    ('datetime', 'datetime', 'Date Time', 'fields.DateTime', 'DATETIME',
        datetime, None),
    ('time', 'time', 'Time', 'fields.Time', 'TIME', time, None),
    ('timestamp', 'timestamp', 'Timestamp', 'fields.Timestamp', 'TIMESTAMP',
        datetime, None),
    ('timedelta', 'timedelta', 'Time Interval', 'fields.TimeDelta', 'INTERVAL',
        relativedelta, None),
    ('icon', 'char', 'Icon', 'fields.Char', 'VARCHAR', str, None),
    ('image', 'binary', 'Image', 'fields.Binary', 'BLOB', bytes, bytearray),
    ('binary', 'binary', 'File', 'fields.Binary', 'BLOB', bytes, bytearray),
    ('reference', 'reference', 'Reference', 'fields.Reference', 'VARCHAR', str,
        None),
    ('selection', 'selection', 'Selection', 'fields.Selection', 'VARCHAR', str,
        None),
    ]

FIELD_TYPE_SELECTION = [(x[0], x[2]) for x in FIELD_TYPES]
FIELD_TYPE_SQL = dict([(x[0], x[4]) for x in FIELD_TYPES])
FIELD_TYPE_CLASS = dict([(x[0], x[3]) for x in FIELD_TYPES])
FIELD_TYPE_PYTHON = dict([(x[0], x[5]) for x in FIELD_TYPES])
FIELD_TYPE_TRYTON = dict([(x[0], x[1]) for x in FIELD_TYPES])
FIELD_TYPE_CAST = dict([(x[0], x[6]) for x in FIELD_TYPES])

VALID_FIRST_SYMBOLS = 'abcdefghijklmnopqrstuvwxyz'
VALID_NEXT_SYMBOLS = '_0123456789'
VALID_SYMBOLS = VALID_FIRST_SYMBOLS + VALID_NEXT_SYMBOLS

BLOCKSIZE = 65536

if config.getboolean('lims_interface', 'filestore', default=False):
    file_id = 'origin_file_id'
    store_prefix = config.get('lims_interface', 'store_prefix', default=None)
else:
    file_id = None
    store_prefix = None


def convert_to_symbol(text):
    if not text:
        return 'x'
    text = unidecode.unidecode(text)
    text = text.lower()
    first = text[0]
    symbol = first
    if first not in VALID_FIRST_SYMBOLS:
        symbol = '_'
        if symbol in VALID_SYMBOLS:
            symbol += first

    for x in text[1:]:
        if x not in VALID_SYMBOLS:
            if symbol[-1] != '_':
                symbol += '_'
        else:
            symbol += x
    return symbol


def str2date(value, lang=None):
    Lang = Pool().get('ir.lang')
    if lang is None:
        lang = Lang.get()
    try:
        return datetime.strptime(value, lang.date)
    except Exception:
        return datetime.strptime(value, '%Y/%m/%d')


def get_model_resource(model_name, value, field_name):
    if value.startswith('='):
        return None
    Model = Pool().get(model_name)
    rec_name = Model._rec_name
    if rec_name not in Model._fields:
        rec_name = 'id'
        try:
            value = int(value)
        except Exception:
            raise UserError(gettext(
                'lims_interface.invalid_default_value_many2one_id',
                name=field_name))
    resource = Model.search([rec_name, '=', value])
    if not resource or len(resource) > 1:
        if 'code' in Model._fields:
            resource = Model.search(['code', '=', value])
        if not resource or len(resource) > 1:
            raise UserError(gettext(
                'lims_interface.invalid_default_value_many2one',
                name=field_name))
    return resource


class Interface(Workflow, ModelSQL, ModelView):
    'Interface'
    __name__ = 'lims.interface'

    _controller_states = {
        'invisible': Eval('kind') != 'controller',
        'required': Eval('kind') == 'controller',
        'readonly': Eval('state') != 'draft'
        }
    _template_states = {
        'invisible': Eval('kind') != 'template',
        'required': Eval('kind') == 'template',
        'readonly': Eval('state') != 'draft',
        }

    name = fields.Char('Name', required=True)
    revision = fields.Integer('Revision', required=True, readonly=True)
    language = fields.Many2One('ir.lang', 'Language',
        domain=[('translatable', '=', True)],
        states={'readonly': Eval('state') != 'draft'})
    kind = fields.Selection([
        ('template', 'Template'),
        ('controller', 'Controller'),
        ], 'Kind', required=True,
        states={'readonly': Eval('state') != 'draft'})
    state = fields.Selection([
        ('draft', 'Draft'),
        ('active', 'Active'),
        ('cancelled', 'Cancelled'),
        ], 'State', readonly=True, required=True)
    controller_name = fields.Selection([(None, '')], 'Controller Name',
        sort=False, states=_controller_states)
    columns = fields.One2Many('lims.interface.column', 'interface', 'Columns',
        states={
            'readonly': Eval('state') != 'draft',
            'invisible': Eval('kind') != 'template',
            })
    grouped_repetitions = fields.One2Many('lims.interface.grouped_repetition',
        'interface', 'Grouped repetitions',
        states={
            'readonly': Eval('state') != 'draft',
            'invisible': Eval('kind') != 'template',
            })
    views = fields.One2Many('lims.interface.view', 'interface', 'Views',
        states={
            'readonly': Eval('state') != 'draft',
            'invisible': Eval('kind') != 'template',
            })
    table = fields.Many2One('lims.interface.table', 'Table', readonly=True)
    template_type = fields.Selection([
        (None, ''),
        ('excel', 'Excel'),
        ('csv', 'Comma Separated Values'),
        ('txt', 'Text File'),
        ], 'Template Type',
        states=_template_states)
    first_row = fields.Integer('First Row', states=_template_states)
    field_separator = fields.Selection([
        ('comma', 'Comma (,)'),
        ('colon', 'Colon (:)'),
        ('semicolon', 'Semicolon (;)'),
        ('tab', 'Tab'),
        ('space', 'Space'),
        ('other', 'Other'),
        ], 'Field separator',
        states={
            'required': And(Eval('kind') == 'template',
                Eval('template_type') == 'csv'),
            'invisible': Or(Eval('kind') != 'template',
                Eval('template_type') != 'csv'),
            'readonly': Eval('state') != 'draft',
            })
    field_separator_other = fields.Char('Other',
        states={
            'required': And(Eval('template_type') == 'csv',
                Eval('field_separator') == 'other'),
            'invisible': Or(Eval('template_type') != 'csv',
                Eval('field_separator') != 'other'),
            'readonly': Eval('state') != 'draft',
            })
    fraction_field = fields.Many2One('lims.interface.column',
        'Fraction field',
        domain=[('interface', '=', Eval('id')), ('group', '=', None)],
        states={'readonly': Eval('state') != 'draft'})
    analysis_field = fields.Many2One('lims.interface.column',
        'Analysis field',
        domain=[('interface', '=', Eval('id')), ('group', '=', None)],
        states={'readonly': Eval('state') != 'draft'})
    method_field = fields.Many2One('lims.interface.column',
        'Method field',
        domain=[('interface', '=', Eval('id')), ('group', '=', None)],
        states={'readonly': Eval('state') != 'draft'})
    repetition_field = fields.Many2One('lims.interface.column',
        'Repetition field',
        domain=[('interface', '=', Eval('id')), ('group', '=', None)],
        states={'readonly': Eval('state') != 'draft'})
    charset = fields.Selection([
        (None, ''),
        ('utf-8', 'UTF-8'),
        ('iso-8859-1', 'ISO-8859-1'),
        ], 'Charset')
    form_col = fields.Integer('Form Columns')

    del _controller_states, _template_states

    @classmethod
    def __setup__(cls):
        super().__setup__()
        cls._transitions |= set((
            ('draft', 'active'),
            ('active', 'cancelled'),
            ('active', 'draft'),
            ))
        cls._buttons.update({
            'draft': {
                'icon': 'tryton-undo',
                'invisible': Eval('state') != 'active',
                },
            'activate': {
                'icon': 'tryton-ok',
                'invisible': Eval('state') != 'draft',
                },
            'copy_columns': {
                'invisible': Eval('state') != 'draft',
                },
            'show_view': {
                'invisible': Eval('state') != 'active',
                },
            })

    @classmethod
    def __register__(cls, module_name):
        cursor = Transaction().connection.cursor()
        sql_table = cls.__table__()
        super().__register__(module_name)
        cursor.execute(*sql_table.update(
                [sql_table.state], ['cancelled'],
                where=sql_table.state == 'canceled'))

    @classmethod
    def view_attributes(cls):
        return [
            ('//notebook/page[@id="columns"]', 'states', {
                    'invisible': Eval('kind') != 'template',
                    }),
            ('//notebook/page[@id="fields"]', 'states', {
                    'invisible': Eval('kind') != 'template',
                    }),
            ('//notebook/page[@id="template"]', 'states', {
                    'invisible': Eval('kind') != 'template',
                    }),
            ('//notebook/page[@id="controller"]', 'states', {
                    'invisible': Eval('kind') != 'controller',
                    }),
            ]

    @staticmethod
    def default_revision():
        return 0

    @staticmethod
    def default_kind():
        return 'template'

    @staticmethod
    def default_charset():
        return 'utf-8'

    @staticmethod
    def default_field_separator():
        return 'semicolon'

    @staticmethod
    def default_state():
        return 'draft'

    @property
    def data_table_name(self):
        return ('lims.interface.table.data.%d.%d' % (self.id or 0,
                self.revision)).replace('.', '_')

    @classmethod
    @ModelView.button
    @Workflow.transition('draft')
    def draft(cls, interfaces):
        pass

    @classmethod
    @ModelView.button
    @Workflow.transition('active')
    def activate(cls, interfaces):
        pool = Pool()
        Table = pool.get('lims.interface.table')
        Field = pool.get('lims.interface.table.field')
        GroupedField = pool.get('lims.interface.table.grouped_field')

        def get_inputs(formula):
            if not formula:
                return
            parser = formulas.Parser()
            ast = parser.ast(formula)[1].compile()
            return (' '.join([x for x in ast.inputs])).lower()

        for interface in interfaces:
            # interface.check_formulas()
            # interface.check_icons()

            interface.revision += 1
            table = Table()
            table.name = interface.data_table_name
            fields = {}
            grouped_fields = []

            pos_group = 0
            for grouped_repetition in interface.grouped_repetitions:
                reps = (grouped_repetition.repetitions or 1) + 1
                for rep in range(1, reps):
                    pos = 0
                    for column in interface.columns:
                        if not column.type_:
                            continue
                        if (column.group and
                                column.group != grouped_repetition.group):
                            continue
                        pos += 1
                        if not column.group:
                            if rep == 1 and grouped_repetition.group == 1:
                                position = pos * 1000
                                fields[position] = Field(
                                    name=column.alias,
                                    string=column.name,
                                    type=column.type_,
                                    help=column.expression,
                                    domain=column.domain,
                                    transfer_field=column.transfer_field,
                                    related_line_field=(
                                        column.related_line_field),
                                    related_model=column.related_model,
                                    selection=column.selection,
                                    formula=(column.expression if
                                        column.expression and
                                        column.expression.startswith('=') else
                                        None),
                                    inputs=(get_inputs(column.expression) if
                                        column.expression and
                                        column.expression.startswith('=') else
                                        None),
                                    required=column.required,
                                    readonly=column.readonly,
                                    invisible=column.invisible,
                                    digits=column.digits,
                                    default_width=column.default_width,
                                    colspan=column.colspan,
                                    related_group=column.related_group,
                                    group_name=column.group_name,
                                    group_colspan=column.group_colspan,
                                    group_col=column.group_col,
                                    )
                            continue

                        else:  # column.grouped
                            if rep == 1:
                                position = pos * 1000 + column.group * 100000
                                pos_group = position
                                expression = (column.expression and
                                    column.expression.replace(
                                        '_XX', '').replace(
                                        'XX', '%s' % rep))
                                grouped_fields.append(GroupedField(
                                    name=column.alias,
                                    string=column.name,
                                    type=column.type_,
                                    help=expression,
                                    domain=column.domain,
                                    related_model=column.related_model,
                                    selection=column.selection,
                                    formula=(expression if
                                        expression and
                                        expression.startswith('=') else
                                        None),
                                    inputs=(get_inputs(expression) if
                                        expression and
                                        expression.startswith('=') else None),
                                    required=column.required,
                                    readonly=column.readonly,
                                    invisible=column.invisible,
                                    digits=column.digits,
                                    group=column.group,
                                    default_width=column.default_width,
                                    ))
                            else:
                                pos_group += 1
                                position = pos_group

                            expression = (column.expression and
                                column.expression.replace(
                                    '_XX', '_%s' % rep).replace(
                                    'XX', '%s' % rep))
                            fields[position] = Field(
                                name='%s_%s' % (column.alias, str(rep)),
                                string='%s (%s)' % (column.name, str(rep)),
                                type=column.type_,
                                help=expression,
                                domain=column.domain,
                                transfer_field=column.transfer_field,
                                related_line_field=column.related_line_field,
                                related_model=column.related_model,
                                selection=column.selection,
                                formula=(expression if expression and
                                    expression.startswith('=') else None),
                                inputs=(get_inputs(expression)
                                    if expression and
                                    expression.startswith('=') else None),
                                required=column.required,
                                readonly=column.readonly,
                                invisible=column.invisible,
                                digits=column.digits,
                                group=column.group,
                                related_group=column.related_group,
                                default_width=column.default_width,
                                colspan=column.colspan,
                                group_name=column.group_name,
                                group_colspan=column.group_colspan,
                                group_col=column.group_col,
                                )
            else:
                if not fields:
                    pos = 0
                    for column in interface.columns:
                        if not column.type_:
                            continue
                        pos += 1
                        position = pos * 1000
                        fields[position] = Field(
                            name=column.alias,
                            string=column.name,
                            type=column.type_,
                            help=column.expression,
                            domain=column.domain,
                            transfer_field=column.transfer_field,
                            related_line_field=column.related_line_field,
                            related_model=column.related_model,
                            selection=column.selection,
                            formula=(column.expression if
                                column.expression and
                                column.expression.startswith('=') else
                                None),
                            inputs=(get_inputs(column.expression) if
                                column.expression and
                                column.expression.startswith('=') else
                                None),
                            required=column.required,
                            readonly=column.readonly,
                            invisible=column.invisible,
                            digits=column.digits,
                            related_group=column.related_group,
                            default_width=column.default_width,
                            colspan=column.colspan,
                            group_name=column.group_name,
                            group_colspan=column.group_colspan,
                            group_col=column.group_col,
                            )

            table.fields_ = [fields[x] for x in sorted(fields.keys())]
            table.grouped_fields_ = grouped_fields
            table.create_table()
            table.save()
            interface.table = table

        cls.save(interfaces)
        cls.set_views(interfaces)
        cls.set_grouped_views(interfaces)

    @classmethod
    def set_views(cls, interfaces):
        View = Pool().get('lims.interface.table.view')

        to_delete = []
        to_save = []
        for interface in interfaces:
            if interface.table:
                to_delete += View.search([
                    ('table', '=', interface.table),
                    ])

            for view_type in ['tree', 'form']:
                view = View()
                view.table = interface.table
                view_info = getattr(interface, 'get_%s_view' % view_type)()
                view.arch = view_info['arch']
                view.type = view_info['type']
                to_save.append(view)

        if to_delete:
            View.delete(to_delete)
        if to_save:
            View.save(to_save)

    def get_tree_view(self):
        fields = self._get_fields_tree_view()
        xml = ('<?xml version="1.0"?>\n'
            '<tree editable="1">\n'
            '%s\n'
            '</tree>') % ('\n'.join(fields))
        return {
            'type': 'tree',
            'arch': xml,
            }

    def _get_fields_tree_view(self):
        fields = []
        current_icon = None
        for line in self.table.fields_:
            if line.type in ('datetime', 'timestamp'):
                fields.append('<field name="%s" widget="date"/>' %
                    line.name)
                fields.append('<field name="%s" widget="time"/>' %
                    line.name)
                continue

            if line.type == 'icon':
                current_icon = line.name
                continue

            attributes = []
            if current_icon:
                attributes.append('icon="%s"' % current_icon)
                current_icon = None
            if line.type == 'image':
                attributes.append('widget="image"')
            if line.default_width:
                attributes.append('width="%s"' % line.default_width)

            fields.append('<field name="%s" %s/>' % (line.name,
                    ' '.join(attributes)))

        fields.append('<field name="annulled"/>')
        fields.append('<field name="notebook_line"/>')
        return fields

    def get_form_view(self):
        fields = self._get_fields_form_view()
        xml = ('<?xml version="1.0"?>\n'
            '<form%s>\n'
            '%s\n'
            '</form>') % (' col="%s"' % self.form_col if self.form_col else '',
                '\n'.join(fields))
        return {
            'type': 'form',
            'arch': xml,
            }

    def _get_fields_form_view(self):
        fields = []
        groups = []
        groups_count = 0
        current_group = None

        for line in self.table.fields_:
            groups_count = max(groups_count, line.group or 0)
            if line.group:
                continue

            if line.group_name:
                if line.group_name != current_group:
                    if current_group:
                        fields.append('</group>')
            else:
                if current_group:
                    fields.append('</group>')
                    current_group = None

            if line.related_group:
                for group in range(1, line.related_group + 1):
                    if group not in groups:
                        field_colspan = 4
                        for rep in self.grouped_repetitions:
                            if rep.group == group:
                                field_colspan = rep.colspan or field_colspan
                        fields.append(
                            '<field name="group_%s" colspan="%s"/>' % (
                                group, field_colspan))
                        groups.append(group)

            if line.group_name:
                if line.group_name != current_group:
                    current_group = line.group_name
                    fields.append('<group id="%s" string="%s" \
                        colspan="%s" col="%s">' % (
                            line.group_name, line.group_name,
                            line.group_colspan, line.group_col))

            if line.type != 'multiline' and not line.invisible:
                fields.append('<label name="%s"/>' % line.name)

            if line.type in ('datetime', 'timestamp'):
                fields.append('<group col="2">'
                    '<field name="%s" widget="date"/>'
                    '<field name="%s" widget="time"/>'
                    '</group>' % (line.name, line.name))
                continue

            if line.type == 'icon':
                fields.append('<image name="%s"/>' % line.name)
                continue

            attributes = []
            if line.type == 'image':
                attributes.append('widget="image"')
            if line.colspan:
                attributes.append('colspan="%s"' % line.colspan)
            if line.invisible:
                attributes.append('invisible="1"')

            fields.append('<field name="%s" %s/>' % (line.name,
                    ' '.join(attributes)))

        for i in range(1, groups_count - 1):
            if i not in groups:
                fields.append('<field name="group_%s" colspan="4"/>' % (
                    i + 1, ))

        if current_group:
            fields.append('</group>')

        fields.append('<label name="notebook_line"/>')
        fields.append('<field name="notebook_line"/>')

        return fields

    @classmethod
    def set_grouped_views(cls, interfaces):
        View = Pool().get('lims.interface.table.grouped_view')

        to_delete = []
        to_save = []
        for interface in interfaces:
            if interface.table:
                to_delete += View.search([
                    ('table', '=', interface.table),
                    ])

            for repetition_group in interface.grouped_repetitions:
                for view_type in ['tree', 'form']:
                    view = View()
                    view.table = interface.table
                    view_info = getattr(interface,
                        'get_%s_grouped_view' % view_type)(
                            repetition_group.group)
                    view.arch = view_info['arch']
                    view.type = view_info['type']
                    view.group = repetition_group.group
                    to_save.append(view)

        if to_delete:
            View.delete(to_delete)
        if to_save:
            View.save(to_save)

    def get_tree_grouped_view(self, group):
        fields = self._get_fields_tree_grouped_view(group)
        xml = ('<?xml version="1.0"?>\n'
            '<tree editable="1">\n'
            '%s\n'
            '</tree>') % ('\n'.join(fields))
        return {
            'type': 'tree',
            'arch': xml,
            }

    def _get_fields_tree_grouped_view(self, group):
        fields = []
        fields.append('<field name="iteration" width="60"/>')
        current_icon = None
        for line in self.table.grouped_fields_:
            if line.group != group:
                continue
            if line.type in ('datetime', 'timestamp'):
                fields.append('<field name="%s" widget="date"/>' %
                    line.name)
                fields.append('<field name="%s" widget="time"/>' %
                    line.name)
                continue

            if line.type == 'icon':
                current_icon = line.name
                continue

            attributes = []
            if current_icon:
                attributes.append('icon="%s"' % current_icon)
                current_icon = None
            if line.type == 'image':
                attributes.append('widget="image"')
            if line.default_width:
                attributes.append('width="%s"' % line.default_width)

            fields.append('<field name="%s" %s/>' % (line.name,
                    ' '.join(attributes)))

        fields.append('<field name="data"/>')

        return fields

    def get_form_grouped_view(self, group):
        fields = self._get_fields_form_grouped_view(group)
        xml = ('<?xml version="1.0"?>\n'
            '<form>\n'
            '%s\n'
            '</form>') % '\n'.join(fields)
        return {
            'type': 'form',
            'arch': xml,
            }

    def _get_fields_form_grouped_view(self, group):
        fields = []
        fields.append('<label name="iteration"/>')
        fields.append('<field name="iteration"/>')
        for line in self.table.grouped_fields_:
            if line.group != group:
                continue
            fields.append('<label name="%s"/>' % line.name)
            if line.type in ('datetime', 'timestamp'):
                fields.append('<group col="2">'
                    '<field name="%s" widget="date"/>'
                    '<field name="%s" widget="time"/>'
                    '</group>' % (line.name, line.name))
                continue

            if line.type == 'icon':
                fields.append('<image name="%s"/>' % line.name)
                continue

            attributes = []
            if line.type == 'image':
                attributes.append('widget="image"')

            fields.append('<field name="%s" %s/>' % (line.name,
                    ' '.join(attributes)))

        return fields

    @classmethod
    @ModelView.button_action('lims_interface.wiz_interface_copy_column')
    def copy_columns(cls, interfaces):
        pass

    @classmethod
    @ModelView.button_action('lims_interface.wiz_interface_show_view')
    def show_view(cls, interfaces):
        pass


class Column(sequence_ordered(), ModelSQL, ModelView):
    'Column'
    __name__ = 'lims.interface.column'
    _states = {
        'readonly': Eval('interface_state') != 'draft',
        }

    interface = fields.Many2One('lims.interface', 'Interface',
        required=True, ondelete='CASCADE', states=_states)
    interface_state = fields.Function(fields.Selection('get_interface_states',
        'Interface State'), 'on_change_with_interface_state')
    name = fields.Char('Name', required=True, states=_states)
    alias = fields.Char('Alias', required=True, states=_states)
    evaluation_order = fields.Integer('Evaluation order', states=_states)
    expression = fields.Char('Formula',
        states={'readonly': _states['readonly'] | Bool(Eval('default_value'))},
        help=('In grouped columns the suffix _XX will be replaced by the '
            'corresponding repetition'))
    expression_icon = fields.Function(fields.Char('Expression Icon'),
        'on_change_with_expression_icon')
    domain = fields.Char('Domain Value', states=_states)
    type_ = fields.Selection([
        (None, ''),
        ('char', 'Text (single-line)'),
        ('multiline', 'Text (multi-line)'),
        ('integer', 'Integer'),
        ('float', 'Float'),
        ('numeric', 'Numeric'),
        ('boolean', 'Boolean'),
        ('many2one', 'Link To Kalenis'),
        ('date', 'Date'),
        ('datetime', 'Date Time'),
        ('time', 'Time'),
        ('timestamp', 'Timestamp'),
        ('timedelta', 'Time Interval'),
        ('icon', 'Icon'),
        ('image', 'Image'),
        ('binary', 'File'),
        ('reference', 'Reference'),
        ('selection', 'Selection'),
        ], 'Field Type', states=_states)
    related_model = fields.Many2One('ir.model', 'Related Model',
        states={
            'required': Eval('type_') == 'many2one',
            'invisible': Eval('type_') != 'many2one',
            'readonly': _states['readonly'],
            })
    selection = fields.Text('Selection',
        states={
            'required': Eval('type_') == 'selection',
            'invisible': Eval('type_') != 'selection',
            'readonly': _states['readonly'],
            },
        help='A couple of key and label separated by ":" per line.')
    default_value = fields.Char('Default value',
        states={'readonly': _states['readonly'] | Bool(Eval('expression'))})
    readonly = fields.Boolean('Read only', states=_states)
    invisible = fields.Boolean('Invisible', states=_states)
    required = fields.Boolean('Required', states=_states)
    digits = fields.Integer('Digits',
        states={
            'required': Eval('type_').in_(['float', 'numeric']),
            'invisible': ~Eval('type_').in_(['float', 'numeric']),
            'readonly': _states['readonly'],
            })
    source_start = fields.Integer('Field start',
        states={
            'required': Eval('_parent_interface', {}).get(
                'template_type') == 'txt',
            'invisible': Eval('_parent_interface', {}).get(
                'template_type') != 'txt',
            'readonly': _states['readonly'],
            })
    source_end = fields.Integer('Field end',
        states={
            'required': Eval('_parent_interface', {}).get(
                'template_type') == 'txt',
            'invisible': Eval('_parent_interface', {}).get(
                'template_type') != 'txt',
            'readonly': _states['readonly'],
            })
    source_column = fields.Integer('Column',
        states={
            'readonly': _states['readonly'] | Bool(Eval('default_value')),
            'invisible': Eval('_parent_interface', {}).get(
                'template_type') == 'txt',
            },
        help='Mapped column in source file')
    singleton = fields.Boolean('Is a singleton value',
        states={
            'readonly': _states['readonly'] | Bool(Eval('default_value')),
            'invisible': Eval('_parent_interface', {}).get(
                'template_type') == 'txt',
            },
        help='Is a fixed value (column:row) in source file')
    source_row = fields.Integer('Row',
        states={
            'required': Bool(Eval('singleton')),
            'invisible': Or(Not(Eval('singleton')),
                Eval('_parent_interface', {}).get('template_type') == 'txt'),
            'readonly': _states['readonly'],
            })
    transfer_field = fields.Boolean('Transfer field', states=_states,
        help='Check if value have to be transferred to notebook line')
    related_line_field = fields.Many2One('ir.model.field', 'Related field',
        domain=[('model.model', '=', 'lims.notebook.line')],
        states={
            'required': Bool(Eval('transfer_field')),
            'invisible': Not(Eval('transfer_field')),
            'readonly': _states['readonly'],
            })
    group = fields.Integer('Group', states=_states)
    related_group = fields.Integer('Related Group', states=_states)
    default_width = fields.Integer('Default Width', states=_states)
    colspan = fields.Integer('Colspan', states=_states)
    group_name = fields.Char('Group Name', states=_states)
    group_colspan = fields.Integer('Group Colspan', states=_states)
    group_col = fields.Integer('Group Col', states=_states)

    del _states

    @classmethod
    def __setup__(cls):
        super().__setup__()
        t = cls.__table__()
        cls._sql_constraints += [
            ('interface_alias_uniq',
                Unique(t, t.interface, sql.Column(t, 'alias')),
                'lims_interface.msg_interface_column_alias_unique')
            ]

    @staticmethod
    def default_digits():
        return 2

    @classmethod
    def default_interface_state(cls):
        return 'draft'

    @classmethod
    def get_interface_states(cls):
        pool = Pool()
        Interface = pool.get('lims.interface')
        return Interface.fields_get(['state'])['state']['selection']

    @fields.depends('default_value', 'expression', 'singleton',
        'source_column')
    def on_change_default_value(self):
        if self.default_value:
            self.expression = None
            self.singleton = False
            self.source_column = None

    @fields.depends('name', 'alias', 'interface',
        '_parent_interface.columns', 'evaluation_order')
    def on_change_name(self):
        if not self.alias:
            self.alias = convert_to_symbol(self.name)
        if not self.evaluation_order and self.interface:
            self.evaluation_order = len(self.interface.columns)

    @fields.depends('interface', '_parent_interface.state')
    def on_change_with_interface_state(self, name=None):
        if self.interface:
            return self.interface.state

    @classmethod
    def validate(cls, columns):
        for column in columns:
            column.check_alias()
            column.check_default_value()
            column.check_domain()
            column.check_selection()

    def check_alias(self):
        for symbol in self.alias:
            if symbol not in VALID_SYMBOLS:
                raise UserError(gettext('lims_interface.invalid_alias',
                    symbol=symbol, name=self.name))

    def check_default_value(self):
        if not self.default_value:
            return
        if self.type_ in [
                'datetime', 'time', 'timestamp', 'timedelta',
                'icon', 'image', 'binary', 'reference',
                ]:
            raise UserError(gettext(
                'lims_interface.invalid_default_value_type',
                name=self.name))
        if self.type_ == 'boolean':
            try:
                int(self.default_value)
            except Exception:
                raise UserError(gettext(
                    'lims_interface.invalid_default_value_boolean',
                    name=self.name))
        elif self.type_ == 'date':
            try:
                str2date(self.default_value, self.interface.language)
            except Exception:
                raise UserError(gettext(
                    'lims_interface.invalid_default_value_date',
                    name=self.name))
        elif self.type_ == 'many2one':
            get_model_resource(
                self.related_model.model, self.default_value, self.name)
        else:
            ftype = FIELD_TYPE_PYTHON[self.type_]
            try:
                ftype(self.default_value)
            except Exception:
                raise UserError(gettext(
                    'lims_interface.invalid_default_value',
                    value=self.default_value, name=self.name))

    def check_domain(self):
        if not self.domain:
            return
        try:
            value = PYSONDecoder().decode(self.domain)
        except Exception:
            raise UserError(gettext(
                'lims_interface.invalid_domain',
                name=self.name))
        if not isinstance(value, list):
            raise UserError(gettext(
                'lims_interface.invalid_domain',
                name=self.name))

    def check_selection(self):
        if self.type_ != 'selection':
            return
        try:
            dict(json.loads(self.get_selection_json()))
        except Exception:
            raise UserError(gettext(
                'lims_interface.invalid_selection',
                name=self.name))

    def get_selection_json(self, name=None):
        db_selection = self.selection or ''
        selection = [[w.strip() for w in v.split(':', 1)]
            for v in db_selection.splitlines() if v]
        return json.dumps(selection, separators=(',', ':'))

    def formula_error(self):
        if not self.expression:
            return
        if not self.expression.startswith('='):
            return
        parser = formulas.Parser()
        try:
            builder = parser.ast(self.expression)[1]
            # Find missing methods:
            # https://github.com/vinci1it2000/formulas/issues/19#issuecomment-429793111
            missing_methods = [k for k, v in builder.dsp.function_nodes.items()
                if v['function'] is formulas.functions.not_implemented]
            if missing_methods:
                # When there are two occurrences of the same missing method,
                # the function name returned looks like this:
                #
                # Sample formula: A(x) + A(y)
                # missing_methods: ['A', 'A<0>']
                #
                # So in the line below we remove the '<0>' suffix
                missing_methods = {x.split('<')[0] for x in missing_methods}
                if len(missing_methods) == 1:
                    msg = 'Unknown method: '
                else:
                    msg = 'Unknown methods: '
                msg += (', '.join(missing_methods))
                return ('error', msg)

            ast = builder.compile()
            missing = (set([x.lower() for x in ast.inputs]) -
                self.previous_formulas())
            if not missing:
                return
            return ('warning', 'Referenced alias "%s" not found. Ensure it is '
                'declared before this formula.' % ', '.join(missing))
        except formulas.errors.FormulaError as error:
            msg = error.msg.replace('\n', ' ')
            if error.args[1:]:
                msg = msg % error.args[1:]
            return ('error', msg)

    def previous_formulas(self):
        res = []
        for formula in self.interface.columns:
            if formula == self:
                break
            res.append(formula.alias)
        return set(res)

    @fields.depends('expression', 'interface', '_parent_interface.columns')
    def on_change_with_expression_icon(self, name=None):
        if not self.expression:
            return ''
        if not self.expression.startswith('='):
            return ''
        error = self.formula_error()
        if not error:
            return 'lims-green'
        if error[0] == 'warning':
            return 'lims-yellow'
        return 'lims-red'


class GroupedRepetition(ModelSQL, ModelView):
    'Grouped Repetition'
    __name__ = 'lims.interface.grouped_repetition'

    interface = fields.Many2One('lims.interface', 'Interface',
        required=True, ondelete='CASCADE')
    group = fields.Integer('Group')
    repetitions = fields.Integer('Repetitions of grouped columns')
    description = fields.Char('Description')
    colspan = fields.Integer('Colspan')


class View(ModelSQL, ModelView):
    'View'
    __name__ = 'lims.interface.view'

    interface = fields.Many2One('lims.interface', 'Interface',
        required=True, ondelete='CASCADE')
    name = fields.Char('Name', required=True)
    columns = fields.One2Many('lims.interface.view.column', 'view', 'Columns',
        context={'interface': Eval('interface')}, depends={'interface'})


class ViewColumn(sequence_ordered(), ModelSQL, ModelView):
    'View Column'
    __name__ = 'lims.interface.view.column'

    view = fields.Many2One('lims.interface.view', 'View',
        required=True, ondelete='CASCADE')
    column = fields.Many2One('lims.interface.column', 'Column', required=True,
        ondelete='CASCADE', domain=['OR',
            ('interface', '=', Eval('context', {}).get('interface', -1)),
            ('interface', '=', Eval('_parent_view', {}).get('interface'))],
        )
    analysis_specific = fields.Boolean('Analysis specific')
    analysis_field = fields.Selection([
        ('code', 'Code'),
        ('description', 'Description'),
        ], 'Analysis field',
        states={'invisible': ~Eval('analysis_specific')})

    @staticmethod
    def default_analysis_specific():
        return False

    @staticmethod
    def default_analysis_field():
        return 'code'

    @classmethod
    def validate(cls, columns):
        super().validate(columns)
        for c in columns:
            c.check_analysis_specific()

    def check_analysis_specific(self):
        if not self.analysis_specific:
            return
        if self.search([
                ('view', '=', self.view.id),
                ('analysis_specific', '=', True),
                ('id', '!=', self.id),
                ]):
            raise UserError(gettext(
                'lims_interface.msg_analysis_specific',
                view=self.view.name))


class CopyInterfaceColumnStart(ModelView):
    'Copy Interface Column'
    __name__ = 'lims.interface.copy_column.start'

    origin_interface = fields.Many2One('lims.interface', 'Origin Interface',
        required=True)


class CopyInterfaceColumn(Wizard):
    'Copy Interface Column'
    __name__ = 'lims.interface.copy_column'

    start = StateTransition()
    ask = StateView('lims.interface.copy_column.start',
        'lims_interface.interface_copy_column_start_view_form', [
            Button('Cancel', 'end', 'tryton-cancel'),
            Button('Copy', 'copy', 'tryton-ok', default=True),
            ])
    copy = StateTransition()

    def transition_start(self):
        Interface = Pool().get('lims.interface')
        interface_id = Transaction().context.get('active_id', None)
        if not interface_id:
            return 'end'
        interface = Interface(interface_id)
        if interface.state != 'draft':
            return 'end'
        return 'ask'

    def transition_copy(self):
        Column = Pool().get('lims.interface.column')

        interface_id = Transaction().context.get('active_id', None)
        count = Column.search_count([('interface', '=', interface_id)])

        new_columns = []
        origin_columns = Column.search([
            ('interface', '=', self.ask.origin_interface.id),
            ], order=[('evaluation_order', 'ASC')])
        for origin in origin_columns:
            count += 1
            column = self._get_column(origin)
            column['interface'] = interface_id
            column['evaluation_order'] = count
            new_columns.append(column)

        if new_columns:
            Column.create(new_columns)
        return 'end'

    def _get_column(self, origin):
        res = {
            'name': origin.name,
            'alias': origin.alias,
            'expression': origin.expression,
            'domain': origin.domain,
            'type_': origin.type_,
            'related_model': (origin.related_model and
                origin.related_model or None),
            'selection': origin.selection,
            'default_value': origin.default_value,
            'readonly': origin.readonly,
            'invisible': origin.invisible,
            'transfer_field': origin.transfer_field,
            'related_line_field': (origin.related_line_field and
                origin.related_line_field.id or None),
            'group': origin.group,
            'digits': origin.digits,
            'related_group': origin.related_group,
            'default_width': origin.default_width,
            'colspan': origin.colspan,
            'group_name': origin.group_name,
            'group_colspan': origin.group_colspan,
            'group_col': origin.group_col,
            }
        return res


class ImportInterfaceColumnStart(ModelView):
    'Import Interface Column Start'
    __name__ = 'lims.interface.import_column.start'

    origin_file = fields.Binary('Origin File', filename='file_name',
        required=True)
    file_name = fields.Char('Name')
    cells = fields.Char('Cells')


class ImportInterfaceColumnMap(ModelView):
    'Import Interface Column Map'
    __name__ = 'lims.interface.import_column.map'

    cells = fields.One2Many(
        'lims.interface.import_column.map.cell', None, 'Cells')


class ImportInterfaceColumnMapCell(ModelView):
    'Import Interface Column Map Cell'
    __name__ = 'lims.interface.import_column.map.cell'

    cell = fields.Char('Cell', required=True)
    name = fields.Char('Name', required=True)
    alias = fields.Char('Alias', required=True)
    expression = fields.Char('Formula')
    group = fields.Integer('Group')

    @fields.depends('name', 'alias')
    def on_change_name(self):
        if not self.alias:
            self.alias = convert_to_symbol(self.name)


class ImportInterfaceColumn(Wizard):
    'Import Interface Column'
    __name__ = 'lims.interface.import_column'

    start = StateTransition()
    ask = StateView('lims.interface.import_column.start',
        'lims_interface.interface_import_column_start_view_form', [
            Button('Cancel', 'end', 'tryton-cancel'),
            Button('Collect', 'collect', 'tryton-forward', default=True),
            ])
    collect = StateTransition()
    map_ = StateView('lims.interface.import_column.map',
        'lims_interface.interface_import_column_map_view_form', [
            Button('Cancel', 'end', 'tryton-cancel'),
            Button('Create', 'create_', 'tryton-forward', default=True),
            ])
    create_ = StateTransition()

    def transition_start(self):
        Interface = Pool().get('lims.interface')
        interface_id = Transaction().context.get('active_id', None)
        if not interface_id:
            return 'end'
        interface = Interface(interface_id)
        if interface.state != 'draft':
            return 'end'
        return 'ask'

    def transition_collect(self):
        def clean_cell(name):
            return name[name.find('!') + 1:]

        tf = tempfile.NamedTemporaryFile(suffix=".xlsx")
        tf.write(self.ask.origin_file)
        xl_model = formulas.ExcelModel().loads(tf.name).finish()
        tf.close()
        cells = {}
        for cell in xl_model.cells:
            if xl_model.cells[cell].inputs:
                for input_ in xl_model.cells[cell].inputs:
                    input_ = clean_cell(input_)
                    if ':' in input_:
                        range_ = re.findall(r'\d+', input_)
                        for row in range(int(range_[0]), int(range_[1]) + 1):
                            if '%s%s' % (input_[:1], row) not in cells.keys():
                                cells['%s%s' % (input_[:1], row)] = {
                                    'input': True}
                    else:
                        if input_ not in cells.keys():
                            cells[input_] = {'input': True}
            if xl_model.cells[cell].builder:
                output = clean_cell(xl_model.cells[cell].output)
                if output not in cells.keys():
                    cells[output] = {'expression':
                        xl_model.cells[cell].builder.match['name']}
                else:
                    cells[output]['expression'] = (
                        xl_model.cells[cell].builder.match['name'])
        self.ask.cells = str(cells)

        return 'map_'

    def default_map_(self, name):
        result_inputs = []
        cells = eval(self.ask.cells)
        for key, value in cells.items():
            result_inputs.append({
                'cell': key,
                'expression': value.get('expression', None),
                })
        return {'cells': sorted(
            result_inputs, key=lambda x: (int(x['cell'][1:]), x['cell'][:1]))}

    def transition_create_(self):
        Column = Pool().get('lims.interface.column')

        interface_id = Transaction().context.get('active_id', None)
        count = Column.search_count([('interface', '=', interface_id)])

        new_columns = []

        for cell in self.map_.cells:
            count += 1
            column = self._get_column(cell)
            column['interface'] = interface_id
            column['evaluation_order'] = count
            new_columns.append(column)

        if new_columns:
            Column.create(new_columns)
        return 'end'

    def _get_column(self, origin):
        expression = None
        if origin.expression:
            expression = self._get_expression(origin.expression)
        res = {
            'name': origin.name,
            'alias': origin.alias,
            'expression': expression,
            'type_': 'float',
            'group': origin.group,
            }
        return res

    def _get_expression(self, expression):
        expression = '=%s' % (expression, )
        parser = formulas.Parser()
        ast = parser.ast(expression)[1].compile()
        inputs = [x for x in ast.inputs]
        for input_ in inputs:
            range_cell = False
            input_value = input_
            if ':' in input_:
                input_value = input_[:input_.find(':')]
                range_cell = True
            for cell in self.map_.cells:
                if input_value == cell.cell:
                    if range_cell:
                        range_ = re.findall(r'\d+', input_)
                        range_alias = []
                        range_count = 1
                        for row in range(int(range_[0]), int(range_[1]) + 1):
                            alias = cell.alias + str(range_count)
                            range_count += 1
                            range_alias.append(alias)
                        expression = expression.replace(
                            input_, ','.join(range_alias))
                    else:
                        expression = expression.replace(input_, cell.alias)
        return expression


class ShowInterfaceViewAsk(ModelView):
    'Show Interface View'
    __name__ = 'lims.interface.show_view.ask'

    notebook_line = fields.Many2One('lims.notebook.line.all_fields',
        'Notebook line')


class ShowInterfaceViewStart(ModelView):
    'Show Interface View'
    __name__ = 'lims.interface.show_view.start'

    data = fields.One2Many('lims.interface.data', None, 'Data')

    @classmethod
    def fields_view_get(cls, view_id=None, view_type='form', level=None):
        result = super().fields_view_get(view_id, view_type, level)
        key = (cls.__name__, view_id, view_type, level)
        cls._fields_view_get_cache.set(key, None)
        return result


class ShowInterfaceView(Wizard):
    'Show Interface View'
    __name__ = 'lims.interface.show_view'

    start_state = 'ask'
    ask = StateView('lims.interface.show_view.ask',
        'lims_interface.interface_show_view_ask_form', [
            Button('Cancel', 'end', 'tryton-cancel'),
            Button('Show', 'start', 'tryton-forward', default=True),
            ])
    start = StateView('lims.interface.show_view.start',
        'lims_interface.interface_show_view_start_form', [
            Button('Close', 'end', 'tryton-close', default=True),
            ])

    def default_start(self, fields):
        pool = Pool()
        Interface = pool.get('lims.interface')

        interface_id = Transaction().context.get('active_id', None)
        if not interface_id:
            return {}

        interface = Interface(interface_id)
        if not interface.table:
            return {}

        fields = interface.table.fields_
        notebook_line_id = (self.ask.notebook_line and
            self.ask.notebook_line.line.id or None)

        record = {
            'notebook_line': notebook_line_id,
            }
        for field in fields:
            if field.group:
                continue
            record[field.name] = None

        grouped_fields = defaultdict(list)
        for field in interface.table.grouped_fields_:
            grouped_fields[field.group].append(field)

        for group, repetition_fields in grouped_fields.items():
            for rep in interface.grouped_repetitions:
                if rep.group == group:
                    reps = (rep.repetitions or 1) + 1
                    break

            group_fields = []
            for rep in range(1, reps):
                grouped_record = {
                    'notebook_line': notebook_line_id,
                    'data': None,
                    'iteration': rep,
                    }
                for field in repetition_fields:
                    grouped_record[field.name] = None
                group_fields.append(grouped_record)
            record['group_%s' % (group, )] = group_fields

        defaults = {
            'data': [record],
            }
        return defaults


class Compilation(Workflow, ModelSQL, ModelView):
    'Interface Compilation'
    __name__ = 'lims.interface.compilation'

    date_time = fields.DateTime('Date', required=True, select=True)
    interface = fields.Many2One('lims.interface', 'Device Interface',
        domain=[('state', '=', 'active')],
        states={'readonly': Eval('state') != 'draft'})
    revision = fields.Integer('Revision',
        states={'readonly': Eval('state') != 'draft'})
    table = fields.Many2One('lims.interface.table', 'Table')
    device = fields.Many2One('lims.lab.device', 'Device',
        domain=[('device_type.non_analytical', '=', False)])
    origins = fields.One2Many('lims.interface.compilation.origin',
       'compilation', 'Origins')
    state = fields.Selection([
        ('draft', 'Draft'),
        ('active', 'Active'),
        ('validated', 'Validated'),
        ('done', 'Done'),
        ('annulled', 'Annulled'),
        ], 'State', readonly=True, required=True)

    @classmethod
    def __setup__(cls):
        super().__setup__()
        cls._order.insert(0, ('date_time', 'DESC'))
        cls._transitions |= set((
            ('draft', 'active'),
            ('draft', 'annulled'),
            ('active', 'draft'),
            ('active', 'validated'),
            ('active', 'annulled'),
            ('validated', 'active'),
            ('validated', 'done'),
            ('validated', 'annulled'),
            ))
        cls._buttons.update({
            'view_data': {
                'invisible': Eval('state') == 'draft',
                'depends': ['state'],
                },
            'draft': {
                'invisible': Eval('state') != 'active',
                'depends': ['state'],
                },
            'activate': {
                'invisible': ~Eval('state').in_(['draft', 'validated']),
                'depends': ['state'],
                },
            'collect': {
                'invisible': Eval('state') != 'active',
                'depends': ['state'],
                },
            'validate_': {
                'invisible': Eval('state') != 'active',
                'depends': ['state'],
                },
            'confirm': {
                'invisible': Eval('state') != 'validated',
                'depends': ['state'],
                },
            })

    def get_rec_name(self, name):
        return self.interface.rec_name + ' / ' + str(self.revision)

    @staticmethod
    def default_date_time():
        return datetime.now()

    @staticmethod
    def default_state():
        return 'draft'

    @fields.depends('interface', 'table', 'revision')
    def on_change_interface(self):
        self.table = self.interface.table if self.interface else None
        self.revision = self.interface.revision if self.interface else None

    @classmethod
    @ModelView.button_action('lims_interface.wiz_compilation_open_data')
    def view_data(cls, compilations):
        pass

    @classmethod
    @ModelView.button
    @Workflow.transition('draft')
    def draft(cls, compilations):
        pass

    @classmethod
    @ModelView.button
    @Workflow.transition('active')
    def activate(cls, compilations):
        pass

    @classmethod
    @ModelView.button_action('lims_interface.act_open_compilation_data')
    def collect(cls, compilations):
        for c in compilations:
            if c.interface.kind == 'template':
                getattr(c, 'collect_%s' % c.interface.template_type)()

    def collect_csv(self, create_new_lines=True):
        pool = Pool()
        Origin = pool.get('lims.interface.compilation.origin')
        Data = pool.get('lims.interface.data')
        NotebookLine = pool.get('lims.notebook.line')

        data_create = []
        data_write = []

        schema, formula_fields = self._get_schema()
        schema_keys = list(schema.keys())
        separator = {
            'comma': ',',
            'colon': ':',
            'semicolon': ';',
            'tab': '\t',
            'space': ' ',
            'other': self.interface.field_separator_other,
            }
        delimiter = separator[self.interface.field_separator]
        first_row = self.interface.first_row - 1
        encoding = self.interface.charset
        with Transaction().set_context(
                lims_interface_table=self.table):
            imported_files = []
            for origin in self.origins:
                if origin.imported:
                    continue
                filedata = io.BytesIO(origin.origin_file)
                wrapper = io.TextIOWrapper(filedata, encoding=encoding)
                try:
                    str_data = io.StringIO(wrapper.read())
                except UnicodeDecodeError:
                    raise UserError(gettext(
                        'lims_interface.invalid_interface_charset'))
                reader = csv.reader(str_data, delimiter=delimiter)
                count = 0
                for row in reader:
                    if count < first_row:
                        count += 1
                        continue
                    if len(row) == 0:
                        continue
                    line = {'compilation': self.id}
                    for k in schema_keys:
                        value = None
                        default_value = schema[k]['default_value']
                        if default_value not in (None, ''):
                            if not create_new_lines:
                                continue
                            if default_value.startswith('='):
                                continue
                            value = default_value
                        else:
                            col = schema[k]['col']
                            if (not row[col - 1] or
                                    not str(row[col - 1]).strip()):
                                line[k] = None
                                continue
                            value = row[col - 1]

                        if schema[k]['type'] == 'integer':
                            line[k] = int(value)
                        elif schema[k]['type'] == 'float':
                            line[k] = float(value)
                        elif schema[k]['type'] == 'numeric':
                            line[k] = Decimal(str(value))
                        elif schema[k]['type'] == 'boolean':
                            line[k] = bool(value)
                        elif schema[k]['type'] == 'date':
                            line[k] = str2date(value, self.interface.language)
                        elif (schema[k]['type'] == 'many2one' and
                                default_value):
                            resource = get_model_resource(
                                schema[k]['model_name'], value,
                                schema[k]['field_name'])
                            line[k] = resource[0].id
                        else:
                            line[k] = str(value)

                    f_fields = sorted(formula_fields.items(),
                        key=lambda x: x[1]['evaluation_order'])
                    for field in f_fields:
                        line[field[0]] = self._get_formula_value(field, line)

                    line['notebook_line'] = self._get_notebook_line(line)
                    if (line['notebook_line']):
                        nl = NotebookLine(line['notebook_line'])
                        for k in schema_keys:
                            default_value = schema[k]['default_value']
                            if (default_value not in (None, '') and
                                    default_value.startswith('=')):
                                path = default_value[1:].split('.')
                                field = path.pop(0)
                                try:
                                    value = getattr(nl, field)
                                    while path:
                                        field = path.pop(0)
                                        value = getattr(value, field)
                                except AttributeError:
                                    value = None
                                line[k] = value
                    line_id = self._get_compilation_line_id(line)
                    if line_id:
                        line['id'] = line_id
                        data_write.append(line)
                    else:
                        data_create.append(line)
                    count += 1
                imported_files.append(origin)

            if imported_files:
                Origin.write(imported_files, {'imported': True})

            if data_create and create_new_lines:
                Data.create(data_create)
            for data in data_write:
                data_line = Data(data['id'])
                del data['id']
                del data['notebook_line']
                del data['compilation']
                Data.write([data_line], data)

    def collect_excel(self, create_new_lines=True):
        pool = Pool()
        Origin = pool.get('lims.interface.compilation.origin')
        Data = pool.get('lims.interface.data')
        NotebookLine = pool.get('lims.notebook.line')

        data_create = []
        data_write = []

        schema, formula_fields = self._get_schema()
        schema_keys = list(schema.keys())
        first_row = self.interface.first_row
        with Transaction().set_context(
                lims_interface_table=self.table):
            imported_files = []
            for origin in self.origins:
                if origin.imported:
                    continue
                filedata = io.BytesIO(origin.origin_file)
                book = load_workbook(filename=filedata, data_only=True)
                sheet = book.active
                max_row = sheet.max_row + 1
                if first_row <= max_row:
                    for i in range(first_row, max_row):
                        line = {'compilation': self.id}
                        for k in schema_keys:
                            value = None
                            default_value = schema[k]['default_value']
                            if default_value not in (None, ''):
                                if not create_new_lines:
                                    continue
                                if default_value.startswith('='):
                                    continue
                                value = default_value
                            else:
                                col = schema[k]['col']
                                row = i
                                if schema[k]['singleton']:
                                    row = schema[k]['row']
                                value = sheet.cell(row=row, column=col).value
                                if value is None:
                                    line[k] = None
                                    continue

                            if schema[k]['type'] == 'integer':
                                line[k] = int(value)
                            elif schema[k]['type'] == 'float':
                                line[k] = float(value)
                            elif schema[k]['type'] == 'numeric':
                                line[k] = Decimal(str(value))
                            elif schema[k]['type'] == 'boolean':
                                line[k] = bool(value)
                            elif schema[k]['type'] == 'date':
                                if default_value:
                                    line[k] = str2date(
                                        value, self.interface.language)
                                else:
                                    if isinstance(value, datetime):
                                        line[k] = value
                                    else:
                                        line[k] = None
                            elif (schema[k]['type'] == 'many2one' and
                                    default_value):
                                resource = get_model_resource(
                                    schema[k]['model_name'], value,
                                    schema[k]['field_name'])
                                line[k] = resource[0].id
                            else:
                                line[k] = str(value)

                        f_fields = sorted(formula_fields.items(),
                            key=lambda x: x[1]['evaluation_order'])
                        for field in f_fields:
                            line[field[0]] = self._get_formula_value(
                                field, line)

                        line['notebook_line'] = self._get_notebook_line(line)
                        if (line['notebook_line']):
                            nl = NotebookLine(line['notebook_line'])
                            for k in schema_keys:
                                default_value = schema[k]['default_value']
                                if (default_value not in (None, '') and
                                        default_value.startswith('=')):
                                    path = default_value[1:].split('.')
                                    field = path.pop(0)
                                    try:
                                        value = getattr(nl, field)
                                        while path:
                                            field = path.pop(0)
                                            value = getattr(value, field)
                                    except AttributeError:
                                        value = None
                                    line[k] = value
                        line_id = self._get_compilation_line_id(line)
                        if line_id:
                            line['id'] = line_id
                            data_write.append(line)
                        else:
                            data_create.append(line)
                imported_files.append(origin)

            if imported_files:
                Origin.write(imported_files, {'imported': True})

            if data_create and create_new_lines:
                Data.create(data_create)
            for data in data_write:
                data_line = Data(data['id'])
                del data['id']
                del data['notebook_line']
                del data['compilation']
                Data.write([data_line], data)

    def collect_txt(self, create_new_lines=True):
        return

    def _get_schema(self):
        schema = {}
        formula_fields = {}
        for column in self.interface.columns:
            if column.source_column or column.default_value:
                schema[column.alias] = {
                    'col': column.source_column,
                    'type': column.type_,
                    'singleton': False,
                    'default_value': None,
                    }
                if column.singleton:
                    schema[column.alias].update({
                        'singleton': True,
                        'row': column.source_row,
                        })
                if column.default_value is not None:
                    schema[column.alias].update({
                        'default_value': column.default_value,
                        })
                    if column.type_ == 'many2one':
                        schema[column.alias].update({
                            'field_name': column.name,
                            'model_name': column.related_model.model,
                            })
            if column.expression:
                formula_fields[column.alias] = {
                    'type': column.type_,
                    'formula': column.expression,
                    'evaluation_order': column.evaluation_order,
                    }
        return schema, formula_fields

    def _get_formula_value(self, field, line):
        parser = formulas.Parser()
        ast = parser.ast(field[1]['formula'])[1].compile()
        inputs = (' '.join([x for x in ast.inputs])).lower().split()
        inputs = [line[x] for x in inputs]
        try:
            value = ast(*inputs)
        except schedula.utils.exc.DispatcherError as e:
            raise UserError(e.args[0] % e.args[1:])

        if isinstance(value, list):
            value = str(value)
        elif not isinstance(value, (str, int, float, Decimal, type(None))):
            value = value.tolist()
        if isinstance(value, formulas.tokens.operand.XlError):
            value = None
        elif isinstance(value, list):
            for x in chain(*value):
                if isinstance(x, formulas.tokens.operand.XlError):
                    value = None
        return value

    def _get_notebook_line(self, line):
        pool = Pool()
        NotebookLine = pool.get('lims.notebook.line')

        fraction_field = self.interface.fraction_field
        analysis_field = self.interface.analysis_field
        repetition_field = self.interface.repetition_field
        if not fraction_field or not analysis_field or not repetition_field:
            return None

        fraction_value = line.get(fraction_field.alias)
        analysis_value = line.get(analysis_field.alias)
        repetition_value = line.get(repetition_field.alias)
        if (fraction_value is None or
                analysis_value is None or
                repetition_value is None):
            return None

        clause = [
            ('notebook.fraction.number', '=', fraction_value),
            ('analysis.code', '=', analysis_value.split(' - ')[0]),
            ('analysis.automatic_acquisition', '=', True),
            ('repetition', '=', repetition_value),
            ('annulled', '=', False),
            ]
        method_field = self.interface.method_field
        if method_field:
            method_value = line.get(method_field.alias)
            if method_value is not None:
                clause.append(
                    ('method.code', '=', method_value.split(' - ')[0]))

        nb_line = NotebookLine.search(clause)
        if nb_line:
            return nb_line[0].id
        return None

    def _get_compilation_line_id(self, line):
        pool = Pool()
        Data = pool.get('lims.interface.data')

        clause = [('compilation', '=', line['compilation'])]
        if line.get('notebook_line'):
            clause.append(('notebook_line', '=', line['notebook_line']))

        else:
            sub_clause = [('id', '=', -1)]
            fraction_field = self.interface.fraction_field
            analysis_field = self.interface.analysis_field
            repetition_field = self.interface.repetition_field
            if fraction_field and analysis_field and repetition_field:

                fraction_value = line.get(fraction_field.alias)
                analysis_value = line.get(analysis_field.alias)
                repetition_value = line.get(repetition_field.alias)

                if (fraction_value is not None and
                        analysis_value is not None and
                        repetition_value is not None):
                    sub_clause = [
                        (fraction_field.alias, '=', fraction_value),
                        (analysis_field.alias, '=', analysis_value),
                        (repetition_field.alias, '=', repetition_value),
                        ]
                    method_field = self.interface.method_field
                    if method_field:
                        method_value = line.get(method_field.alias)
                        if method_value is not None:
                            sub_clause.append(
                                (method_field.alias, '=', method_value))

            clause.extend(sub_clause)
        line = Data.search(clause)
        return line and line[0].id or None

    @classmethod
    @ModelView.button
    @Workflow.transition('validated')
    def validate_(cls, compilations):
        cls.check_required_fields(compilations)

    @classmethod
    def check_required_fields(cls, compilations):
        pool = Pool()
        Field = pool.get('lims.interface.table.field')
        Data = pool.get('lims.interface.data')

        for c in compilations:
            required_columns = Field.search([
                ('table', '=', c.table),
                ('required', '=', True),
                ])
            if not required_columns:
                continue
            with Transaction().set_context(lims_interface_table=c.table):
                lines = Data.search([('compilation', '=', c.id)])
                for line in lines:
                    if line.annulled:
                        continue
                    for column in required_columns:
                        if getattr(line, column.name) is None:
                            raise UserError(gettext(
                                'lims_interface.missing_required_field',
                                field=column.name))

    @classmethod
    @ModelView.button
    @Workflow.transition('done')
    def confirm(cls, compilations):
        pool = Pool()
        ModelData = pool.get('ir.model.data')
        Data = pool.get('lims.interface.data')
        Field = pool.get('lims.interface.table.field')
        NotebookLine = pool.get('lims.notebook.line')

        cls.check_required_fields(compilations)

        avoid_accept_result = Transaction().context.get('avoid_accept_result',
            False)

        now = datetime.now()
        #today = now.date()
        result_modifier_na = ModelData.get_id('lims', 'result_modifier_na')

        for c in compilations:
            fields = {}
            columns = Field.search([
                ('table', '=', c.table.id),
                ('transfer_field', '=', True),
                ])
            for column in columns:
                fields[column.name] = column.related_line_field.name
            if not fields:
                continue
            with Transaction().set_context(lims_interface_table=c.table):
                lines = Data.search([('compilation', '=', c.id)])
                for line in lines:
                    if not cls._allow_confirm_line(line):
                        continue
                    nb_line = line.notebook_line
                    data = {
                        'compilation': c.id,
                        }
                    data_eng = {}
                    for alias, nl_field in fields.items():
                        data[nl_field] = getattr(line, alias, None)
                        if nl_field == 'result' and data[nl_field] is not None:
                            if not nb_line.significant_digits:
                                decimals = nb_line.decimals or 0
                                result = round(float(data[nl_field]), decimals)
                                data[nl_field] = format(result,
                                    '.{}f'.format(decimals))
                        if nl_field == 'result_modifier' and data[nl_field]:
                            data[nl_field] = data[nl_field].id
                        if nl_field == 'literal_result' and data[nl_field]:
                            data_eng[nl_field] = data[nl_field]

                    if line.annulled:
                        data.update({
                            'result_modifier': result_modifier_na,
                            'annulled': True,
                            'annulment_date': now,
                            'report': False,
                            })
                    elif (not avoid_accept_result and
                            nb_line.laboratory.automatic_accept_result):
                        #data['end_date'] = today
                        data['accepted'] = True
                        data['acceptance_date'] = now
                    if data:
                        NotebookLine.write([nb_line], data)
                    if data_eng:
                        with Transaction().set_context(language='en'):
                            NotebookLine.write([nb_line], data_eng)

    @classmethod
    def _allow_confirm_line(cls, line):
        nb_line = line.notebook_line
        if not nb_line:
            return False
        if nb_line.end_date:
            return False
        if nb_line.annulled:
            return False
        return True

    @classmethod
    @ModelView.button
    @Workflow.transition('annulled')
    def annul(cls, compilations):
        pool = Pool()
        Data = pool.get('lims.interface.data')

        for c in compilations:
            with Transaction().set_context(lims_interface_table=c.table):
                lines = Data.search([('compilation', '=', c.id)])
                if not lines:
                    continue
                for line in lines:
                    nb_line = line.notebook_line
                    if not nb_line:
                        continue
                    if not nb_line.annulled:
                        raise UserError(gettext(
                            'lims_interface.msg_line_not_annulled',
                            notebook_line=nb_line.rec_name))

    @classmethod
    def delete(cls, compilations):
        Data = Pool().get('lims.interface.data')
        for c in compilations:
            if c.state == 'done':
                raise UserError(gettext(
                    'lims_interface.delete_done_compilation'))
            with Transaction().set_context(lims_interface_table=c.table):
                lines = Data.search([('compilation', '=', c.id)])
                Data.delete(lines)
        super().delete(compilations)


class CompilationOrigin(ModelSQL, ModelView):
    'Compilation Origin'
    __name__ = 'lims.interface.compilation.origin'

    compilation = fields.Many2One('lims.interface.compilation',
    'Compilation', required=True, ondelete='CASCADE')
    origin_file = fields.Binary("Origin File", filename='file_name',
        file_id=file_id, store_prefix=store_prefix)
    file_name = fields.Char('Name')
    origin_file_id = fields.Char("Origin File ID", readonly=True)
    imported = fields.Boolean('Imported', readonly=True)

    @staticmethod
    def default_imported():
        return False

    @classmethod
    def validate(cls, origins):
        super().validate(origins)
        for origin in origins:
            origin.check_duplicated()

    def check_duplicated(self):
        if len(self.origin_file) == 0:
            return
        for existing in self.compilation.origins:
            if existing.id == self.id:
                continue
            if len(existing.origin_file) == len(self.origin_file):
                existing_f = io.BytesIO(existing.origin_file)
                ef_hash = hashlib.sha1()
                buf = existing_f.read(BLOCKSIZE)
                while len(buf) > 0:
                    ef_hash.update(buf)
                    buf = existing_f.read(BLOCKSIZE)
                new_f = io.BytesIO(self.origin_file)
                nf_hash = hashlib.sha1()
                buf = new_f.read(BLOCKSIZE)
                while len(buf) > 0:
                    nf_hash.update(buf)
                    buf = new_f.read(BLOCKSIZE)

                if ef_hash.hexdigest() == nf_hash.hexdigest():
                    raise UserError(gettext(
                        'lims_interface.duplicated_origin_file',
                        file_name=self.file_name))


class OpenCompilationData(Wizard):
    'Open Compilation Data'
    __name__ = 'lims.interface.compilation.open_data'

    start = StateAction('lims_interface.act_open_compilation_data')

    def do_start(self, action):
        Compilation = Pool().get('lims.interface.compilation')

        context = {
            'lims_interface_compilation': None,
            'lims_interface_table': None,
            'lims_interface_readonly': False,
            }
        domain = [('compilation', '=', None)]

        compilation_id = Transaction().context.get('active_id', None)
        if compilation_id:
            compilation = Compilation(compilation_id)
            readonly = (compilation.state in ('validated', 'done', 'annulled'))
            context['lims_interface_compilation'] = compilation.id
            context['lims_interface_table'] = compilation.table.id
            context['lims_interface_readonly'] = readonly
            domain = [('compilation', '=', compilation.id)]
        action['pyson_context'] = PYSONEncoder().encode(context)
        action['pyson_domain'] = PYSONEncoder().encode(domain)
        return action, {}


class TestFormulaView(ModelView):
    'Test Formula'
    __name__ = 'lims.interface.formula.test'
    expression_column = fields.Many2One('lims.interface.column',
        'Formula Column', required=True,
        domain=[('id', 'in', Eval('expression_column_domain'))])
    expression_column_domain = fields.Function(fields.Many2Many(
        'lims.interface.column', None, None, 'Formula column domain'),
        'on_change_with_expression_column_domain')
    expression = fields.Char('Formula', required=True)
    expression_icon = fields.Function(fields.Char('Expression Icon'),
        'on_change_with_expression_icon')
    variables = fields.One2Many('lims.interface.formula.test.variable',
        None, 'Variables')
    result = fields.Char('Result', readonly=True)

    def formula_error(self):
        if not self.expression:
            return
        if not self.expression.startswith('='):
            return
        parser = formulas.Parser()
        try:
            builder = parser.ast(self.expression)[1]
            # Find missing methods:
            # https://github.com/vinci1it2000/formulas/issues/19#issuecomment-429793111
            missing_methods = [k for k, v in builder.dsp.function_nodes.items()
                if v['function'] is formulas.functions.not_implemented]
            if missing_methods:
                # When there are two occurrences of the same missing method,
                # the function name returned looks like this:
                #
                # Sample formula: A(x) + A(y)
                # missing_methods: ['A', 'A<0>']
                #
                # So in the line below we remove the '<0>' suffix
                missing_methods = {x.split('<')[0] for x in missing_methods}
                if len(missing_methods) == 1:
                    msg = 'Unknown method: '
                else:
                    msg = 'Unknown methods: '
                msg += (', '.join(missing_methods))
                return ('error', msg)

            return

        except formulas.errors.FormulaError as error:
            msg = error.msg.replace('\n', ' ')
            if error.args[1:]:
                msg = msg % error.args[1:]
            return ('error', msg)

    @fields.depends()
    def on_change_with_expression_column_domain(self, name=None):
        InterfaceColumn = Pool().get('lims.interface.column')
        columns = InterfaceColumn.search([
            ('interface', '=', Transaction().context.get('active_id')),
            ('expression', '!=', ''),
        ])
        return [x.id for x in columns]

    @fields.depends('expression_column', 'variables')
    def on_change_expression_column(self):
        self.expression = None
        self.variables = []
        if not self.expression_column:
            return

        self.expression = self.expression_column.expression
        variables = []
        parser = formulas.Parser()
        ast = parser.ast(self.expression)[1].compile()
        inputs = (' '.join([x for x in ast.inputs])).lower().split()
        for input_ in inputs:
            variables.append({'variable': input_})
        self.variables = variables

    @fields.depends('expression_column', 'expression', 'variables')
    def on_change_with_result(self):
        if not self.expression_column or not self.expression:
            return None

        parser = formulas.Parser()
        try:
            ast = parser.ast(self.expression)[1].compile()
        except Exception:
            return None

        expression_inputs = (' '.join([x for x in ast.inputs])).lower().split()
        inputs = []
        for variable in self.variables:
            if variable.variable not in expression_inputs:
                continue
            try:
                input_value = float(variable.value)
            except ValueError:
                input_value = variable.value
            inputs.append(input_value)
        try:
            value = ast(*inputs)
        except schedula.utils.exc.DispatcherError as e:
            raise UserError(e.args[0] % e.args[1:])

        if isinstance(value, list):
            value = str(value)
        elif not isinstance(value, (str, int, float, Decimal, type(None))):
            value = value.tolist()
        if isinstance(value, formulas.tokens.operand.XlError):
            value = None
        elif isinstance(value, list):
            for x in chain(*value):
                if isinstance(x, formulas.tokens.operand.XlError):
                    value = None

        self.result = value
        return self._changed_values.get('result', [])

    @fields.depends('expression')
    def on_change_with_expression_icon(self, name=None):
        if not self.expression:
            return ''
        if not self.expression.startswith('='):
            return ''
        error = self.formula_error()
        if not error:
            return 'lims-green'
        if error[0] == 'warning':
            return 'lims-yellow'
        return 'lims-red'


class TestFormulaViewVariable(ModelView):
    'Test Formula'
    __name__ = 'lims.interface.formula.test.variable'
    variable = fields.Char('Variable', readonly=True)
    value = fields.Text('Value')


class TestFormula(Wizard):
    'Test Formula'
    __name__ = 'lims.interface.formula.test'
    start_state = 'test'
    test = StateView('lims.interface.formula.test',
        'lims_interface.interface_formula_test_view_form',
        [Button('Close', 'end', 'tryton-close', default=True)])

    def default_test(self, fields):
        InterfaceColumn = Pool().get('lims.interface.column')
        columns = InterfaceColumn.search([
            ('interface', '=', Transaction().context.get('active_id')),
            ('expression', '!=', ''),
        ])
        default = {
            'expression_column_domain': [x.id for x in columns],
        }
        return default


class Variable(ModelSQL, ModelView):
    'Interface Variable'
    __name__ = 'lims.interface.variable'

    name = fields.Char('Name', required=True)
    values = fields.One2Many('lims.interface.variable.value',
        'variable', 'Values', required=True)


class VariableValue(ModelSQL, ModelView):
    'Interface Variable Value'
    __name__ = 'lims.interface.variable.value'

    variable = fields.Many2One('lims.interface.variable', 'Variable',
        required=True, ondelete='CASCADE', select=True)
    name = fields.Function(fields.Char('Name'), 'get_name',
        searcher='search_name')
    value = fields.Char('Value', required=True)
    analysis = fields.Many2One('lims.analysis', 'Analysis', required=True)
    product_type = fields.Many2One('lims.product.type', 'Product type')
    matrix = fields.Many2One('lims.matrix', 'Matrix')
    method = fields.Many2One('lims.lab.method', 'Method')

    @classmethod
    def get_name(cls, values, name):
        result = {}
        for v in values:
            result[v.id] = v.variable.name
        return result

    @classmethod
    def search_name(cls, name, clause):
        return [('variable.name',) + tuple(clause[1:])]

    @classmethod
    def get_value(cls, name, analysis, product_type=None, matrix=None,
            method=None):
        if not name or not analysis:
            return None
        clause = [
            ('variable.name', '=', name),
            ('analysis', '=', analysis),
            ('product_type', '=', product_type),
            ('matrix', '=', matrix),
            ('method', '=', method),
            ]
        res = cls.search(clause, limit=1)
        if not res:
            return None
        try:
            val = float(res[0].value)
        except (TypeError, ValueError):
            val = res[0].value
        return val


class Constant(ModelSQL, ModelView):
    'Interface Constant'
    __name__ = 'lims.interface.constant'

    name = fields.Char('Name', required=True)
    parameter1 = fields.Float('Parameter 1')
    parameter2 = fields.Float('Parameter 2')
    parameter3 = fields.Float('Parameter 3')
    value1 = fields.Float('Value 1')
    value2 = fields.Float('Value 2')
    value3 = fields.Float('Value 3')
    value4 = fields.Float('Value 4')
    value5 = fields.Float('Value 5')
    value6 = fields.Float('Value 6')
    value7 = fields.Float('Value 7')
    value8 = fields.Float('Value 8')
    value9 = fields.Float('Value 9')

    @classmethod
    def __setup__(cls):
        super().__setup__()
        cls._order.insert(0, ('name', 'ASC'))
        cls._order.insert(1, ('parameter1', 'ASC'))
        cls._order.insert(2, ('parameter2', 'ASC'))
        cls._order.insert(3, ('parameter3', 'ASC'))

    @classmethod
    def get_constant(cls, name, parameter1=None, parameter2=None,
            parameter3=None, value=None):
        if not name:
            return None
        if not value:
            value = 'value1'

        clause = [('name', '=', name)]
        if parameter1 is not None:
            clause.append(('parameter1', '=', parameter1))
        if parameter2 is not None:
            clause.append(('parameter2', '=', parameter2))
        if parameter3 is not None:
            clause.append(('parameter3', '=', parameter3))
        constant = cls.search(clause)
        if not constant:
            return None

        constant = constant[0]
        if hasattr(constant, value):
            return getattr(constant, value)

        return None
